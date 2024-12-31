import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload
from datetime import datetime, timedelta
from google.oauth2 import service_account
from googleapiclient.discovery import build
import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from datetime import datetime
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from gspread_formatting import *

import config
import googleapi
import pytz
import locale
import time
from gspread_formatting import *
from gspread import authorize
from google.oauth2 import service_account
from openpyxl.utils import get_column_letter

from googleVacationApi import is_file_exists_in_directory

# Google Drive API 설정
SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = config.kakao_json_key_path
credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=credentials)

# 폴더 ID 및 오늘 날짜 파일명
folder_id = config.fileABCD_parent_folder_id
copy_id = '1eK9rdEHm0rovV8H37c7r57e4eEXqUjQPe1LUMlA1QVI'

# 파일 ID 조회 함수
def get_file_id_in_folder(folder_id, file_name):
    try:
        # Google Drive API 쿼리 설정
        query = f"'{folder_id}' in parents and name = '{file_name}' and trashed = false"
        response = drive_service.files().list(
            q=query, spaces='drive', fields='files(id, name)',
            supportsAllDrives=True, includeItemsFromAllDrives=True # Essential!!!!!
        ).execute()
        files = response.get('files', [])

        # 파일이 존재하면 첫 번째 파일의 ID 반환, 없으면 None 반환
        if files:
            file_id = files[0].get('id')
            # print(f"File ID for '{file_name}': {file_id}")
            return file_id
        else:
            print(f"No file named '{file_name}' found in the folder.")
            return None
    except HttpError as error:
        print(f"An error occurred: {error}")
        return None

# 폴더 내 파일이 존재하는지 확인하는 함수
def is_file_exists_in_directory(folder_id, file_name):
    try:
        query = f"'{folder_id}' in parents and name = '{file_name}' and trashed = false"
        response = drive_service.files().list(
            q=query, spaces='drive', fields='files(id, name)',
            supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        files = response.get('files', [])
        return len(files) > 0
    except HttpError as error:
        print(f"An error occurred while checking file existence: {error}")
        return False

# 스프레드시트 생성 함수
def create_spreadsheet(folder_id, file_name):
    template_id = config.investment_table_template_id

    try:
        file_metadata = {
            'name': file_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [folder_id]
        }
        file = drive_service.files().create(body=file_metadata, fields='id', supportsAllDrives=True).execute()
        # print(f"Created spreadsheet with ID: {file.get('id')}")
    except HttpError as error:
        print(f"An error occurred while creating the spreadsheet: {error}")

# 폴더 접근 권한 확인
def check_folder_access(folder_id):
    try:
        drive_service.files().get(fileId=folder_id, supportsAllDrives=True).execute()
        return True
    except HttpError as error:
        print(f"Cannot access folder: {error}")
        return False

# inv_id가 선택된 상황입니다
# ex) 라포랩스 (kv : KV_190 inv : KV_190_5)

# dataframe으로 모두 조회하는 방식임
def load_data(db_1, db_4, db_7):
    db_1_info = db_1
    db_4_info = db_4
    db_7_info = db_7

# file_id에 해당하는 스프레드시트에 DataFrame 반영 함수
def update_basic_with_dataframes(file_id, basic_df):
    # NaN 값을 빈 문자열로 대체
    basic_df = basic_df.fillna('')  

    # '펀드명' 또는 '약정금액'이 포함된 행 제거
    basic_df = basic_df[~basic_df.iloc[:, 0].str.contains("펀드명", na=False)]
    # 카카오 벤처스 포함된 행 제거하기
    basic_df = basic_df[~basic_df.iloc[:, 0].str.contains("카카오벤처스", na=False)]

    client = gspread.authorize(credentials)
    # 스프레드시트 열기
    spreadsheet = client.open_by_key(file_id)
    worksheet = spreadsheet.sheet1  # 첫 번째 시트
    
    # 스프레드시트 초기화
    worksheet.clear()

    # 해당 표의 제목을 설정한다
    worksheet.update('A1', [["1. 운용 중인 투자조합 기본정보"]])

    
    # 첫 번째 행에 열 제목 설정
    headers = [
        '조합명', '결성금액', '투자재원', '결성일', '투자기간 만료일', 
        '존속기간 만료일', '대표펀드매니저', '핵심운용인력', '투자기간'
    ]
    worksheet.append_row(headers, table_range="A3")  # 세 번째 행에 열 제목 추가
    
    # DataFrame의 키 값을 제외하고 데이터만 리스트로 변환하여 추가
    data = basic_df.reset_index(drop=True).values.tolist()  # 인덱스를 초기화하여 키 제거
    worksheet.append_rows(data, table_range="A4")  # 데이터를 네 번째 행부터 추가


def update_invest_with_dataframes(file_id, invest_df):
    client = gspread.authorize(credentials)
    # 스프레드시트 열기
    spreadsheet = client.open_by_key(file_id)
    worksheet = spreadsheet.sheet1  # 첫 번째 시트
    
    # 마지막 데이터가 위치한 행 다음에 데이터 추가 위치 계산
    last_row = len(worksheet.get_all_values()) + 2  # 마지막 행 다음 + 한 줄 띄우기

    # 표 제목 추가
    worksheet.update_cell(last_row, 1, '2. 투자대상 기업정보')
    invest_df = invest_df.fillna('')
    
    # 투자 대상 기업정보 추가
    invest_data = [
        ['(1) 투자대상 기업명', invest_df['회사명'].iloc[0]],
        ['(2) 발굴자', invest_df['발굴자1'].iloc[0], invest_df['발굴자2'].iloc[0], invest_df['발굴자3'].iloc[0]],
        ['(3) 후속/후행 투자 여부', invest_df['INV ID 상태'].iloc[0]],
        ['(4) 투자예정금액', invest_df['투자금액(원화)'].iloc[0]],
        ['(5) 투자예정시기', invest_df['투자 납입일'].iloc[0]]
    ]

    # 데이터 추가
    for i, row_data in enumerate(invest_data):
        title = row_data[0]
        worksheet.update_cell(last_row + i + 1, 1, title)  # 제목을 A열에 추가
        
        # 데이터 값을 B열부터 오른쪽으로 추가
        for j, value in enumerate(row_data[1:]):
            worksheet.update_cell(last_row + i + 1, 2 + j, value)  # 데이터를 B열부터 추가

def update_invest_date(file_id, basic_df, invest_df):
    # 구글 스프레드시트 인증
    credentials = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/spreadsheets'])
    client = gspread.authorize(credentials)
    
    # 날짜 형식 변환
    basic_df['투자기간 만료일'] = pd.to_datetime(basic_df['투자기간 만료일'], format='%Y-%m-%d', errors='coerce')
    invest_df['투자 납입일'] = pd.to_datetime(invest_df['투자 납입일'], format='%Y-%m-%d', errors='coerce')

    # 4번째 행부터 마지막 행까지만 투자기간을 업데이트 (DataFrame 인덱스는 0부터 시작하므로 3부터 슬라이싱)
    for index, row in basic_df.iloc[2:].iterrows():
        # 투자기간 만료일이 투자예정시기보다 크거나 같으면 'O', 아니면 'X'
        if row['투자기간 만료일'] >= invest_df['투자 납입일'].iloc[0]:
            basic_df.at[index, '투자기간'] = 'O'
        else:
            basic_df.at[index, '투자기간'] = 'X'

    # 스프레드시트 열기
    spreadsheet = client.open_by_key(file_id)
    worksheet = spreadsheet.sheet1  # 첫 번째 시트

    # 업데이트할 열 추출 (투자기간 열만)
    투자기간_values = basic_df['투자기간'].values.tolist()
    
    # Google Sheets의 투자기간 열에 값 업데이트
    cell_range = f"I2:I{len(투자기간_values) + 1}"  # 투자기간 열의 위치에 맞춰서 설정
    worksheet.update(cell_range, [[value] for value in 투자기간_values])

def fetch_sheet_as_dataframe(file_id, sheet_name):
    # Google Sheets 파일에 인증하고 열기
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
    spreadsheet = gc.open_by_key(file_id)
    sheet = spreadsheet.worksheet(sheet_name)  # 특정 워크시트 가져오기
    
    # Google Sheets 데이터를 가져와서 데이터프레임으로 변환
    data = sheet.get_all_values()  # 모든 데이터를 리스트로 가져오기
    headers = data.pop(0)  # 첫 번째 행을 헤더로 설정
    df = pd.DataFrame(data, columns=headers)  # 데이터프레임 생성
    
    return df

def update_invest_src(file_id, basic_df):
    global copy_id
    invest_src_id = config.investment_resources_overview_id

    # Google Sheets API 인증 설정
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, scope)
    client = gspread.authorize(creds)

    # 원본 시트 열기
    source_sheet = client.open_by_key(invest_src_id).get_worksheet(1) 

    # 원본 시트에서 전체 데이터 가져오기
    source_data = source_sheet.get_all_values()

    # 복사할 대상 시트 열기
    destination_sheet = client.open_by_key(copy_id).sheet1

    # 대상 시트의 범위에 데이터를 업데이트 (전체 데이터 반영)
    destination_sheet.clear()  # 먼저 대상 시트의 내용을 지웁니다.
    destination_sheet.update('A1', source_data)  # A1 셀부터 데이터를 복사

    dest_data = destination_sheet.get_all_values()

    # 3번째 행에서 빈칸 제외
    third_row_data = [cell for cell in dest_data[2] if cell.strip() != '']

    # 21번째 행에서 '투자집행가능'으로 시작하지 않는 문자열 제외
    twenty_first_row_data = [
        cell for cell in dest_data[20] if cell.strip() != '' and not cell.startswith('투자집행가능')
    ]

    paired_data = list(zip(third_row_data, twenty_first_row_data))

    # 파일 ID에 해당하는 스프레드시트 열기
    target_sheet = client.open_by_key(file_id).sheet1
    target_data = target_sheet.get_all_values()

    # '조합명'이 포함된 행 찾기
    start_row = None
    for idx, row in enumerate(target_data, start=1):  # 1-based indexing
        if '조합명' in row:
            start_row = idx
            break

    if start_row is None:
        print("Error: '조합명'을 찾을 수 없습니다.")
        return
    
    # '조합명' 다음 행부터 업데이트
    update_start_row = start_row + 1
    
    # A열에 조합명 이후로 빈칸찾기
    # 빈칸 전까지 업데이트 진행
    # 조합명 이후 데이터 - 빈칸 전 데이터까지 C열 데이터 업데이트 진행 (범위를 설정해두고)

    # 범위 내에 있는 A열 데이터 추출
    # A열 데이터 내 요소를 하나씩 조회하고
    # paired_data 내 third_row_data에 조회한 요소가 존재하면 매칭된 twenty_first_row_data를 해당 데이터의 C열에 업데이트 한다
    # 범위 내에서 업데이트를 모두 진행한다 
    # A열 데이터 추출 (빈칸 전까지만)
    a_column_data = []
    for row in target_data[update_start_row - 1:]:  # 0-based indexing adjustment
        if len(row) == 0 or row[0].strip() == '':  # 빈칸이면 중지
            break
        a_column_data.append(row[0])

    # 범위 내 데이터 업데이트
    for i, target_name in enumerate(a_column_data):
        # paired_data에서 매칭 데이터 찾기
        update_value = "-"
        for pair in paired_data:
            if target_name == pair[0]:  # A열 데이터와 paired_data의 third_row_data 매칭
                update_value = pair[1]
                break

        # C열에 값 업데이트
        target_cell = f'C{update_start_row + i}'
        target_sheet.update(target_cell, [[update_value]])  # 값은 반드시 2D 리스트로 전달

    return paired_data

def get_basic_info_from_spreadsheet(file_id, sheet_name="Sheet1", target_keyword="조합명"):
    # Google API 자격 증명 설정 (서비스 계정 파일 필요)
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE)
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    
    # 스프레드시트 데이터를 읽어옵니다.
    result = sheet.values().get(spreadsheetId=file_id, range=sheet_name).execute()
    values = result.get('values', [])
    
    # 스프레드시트에서 조합명이 포함된 위치 찾기
    start_index = None
    for i, row in enumerate(values):
        if target_keyword in row:
            start_index = i
            break
    
    # 조합명 이후 데이터를 빈칸 전까지 추출
    if start_index is not None:
        data_rows = []
        for row in values[start_index + 1:]:
            if any(cell.strip() for cell in row):  # 빈 행이 아닌 경우만 추가
                data_rows.append(row)
            else:
                break  # 빈 행을 만나면 중단
    
        # DataFrame으로 변환
        df = pd.DataFrame(data_rows)
        return df
    else:
        print(f"{target_keyword}를 찾을 수 없습니다.")
        return pd.DataFrame()  # 빈 데이터프레임 반환

def update_invest_distribution(file_id, fund_name, inv_id):
    # Google Sheets 파일에 인증하고 열기
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
    spreadsheet = gc.open_by_key(file_id)
    sheet = spreadsheet.sheet1  # 기본 워크시트 선택

    # db7_id 스프레드시트 열기
    db7_id = config.db_7_id
    client = gspread.authorize(credentials)
    db7_sheet = client.open_by_key(db7_id).sheet1  # 첫 번째 시트 사용 (필요 시 시트 이름으로 변경)

    # B열에서 fund_name과 같은 데이터 검색 및 해당 데이터의 D열 반환
    fund_data_in_b = db7_sheet.col_values(2)  # B열의 모든 데이터 가져오기

    # fund_name이 위치한 행 찾기
    row_num = None
    for idx, value in enumerate(fund_data_in_b, start=1):  # 행 번호가 1부터 시작
        if value == fund_name:
            row_num = idx
            break

    # D열의 데이터 반환
    # fund_commitment =>>>> 펀드 약정 금액
    if row_num:
        fund_commitment = db7_sheet.cell(row_num, 5).value  # D열의 데이터 가져오기 (4는 D열에 해당)
        # print(f"펀드 약정 금액 : {fund_commitment}") 
    else:
        print("fund_name을 찾을 수 없습니다.")

    data = sheet.get_all_values()
    last_row = len(data) + 2  # 마지막 데이터 이후 한 줄을 건너뜀

    # '3. 투자재원 배분 결정 : 투자재원 배분기준 지침 참조' - 제목을 먼저 형성하고
    sheet.update(f'A{last_row}', [['3. 투자재원 배분 결정 : 투자재원 배분기준 지침 참조']])
    last_row = last_row + 2

    # 총 3개의 표를 구성합니다
    # 1번째 표 구성 (제목)
    sheet.update(f'A{last_row}', [['(1) 기투자한 조합이 있을 경우 우선 배정 가능']])

    # '(3) 후속/후행 투자 여부' 셀을 찾고 해당 셀의 우측 값을 조회
    try:
        cell = sheet.find('(3) 후속/후행 투자 여부')
        right_cell_value = sheet.cell(cell.row, cell.col + 1).value  # 우측 셀 값 가져오기

        # 조회한 값을 last_row에 업데이트
        sheet.update(f'A{last_row + 1}', [[f'{right_cell_value}']])

    except gspread.exceptions.CellNotFound:
        print("셀 '(3) 후속/후행 투자 여부'를 찾을 수 없습니다.")
    last_row += 3 

    # 2번째 표 구성 (제목 및 헤더)
    sheet.update(f'A{last_row}', [['(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부']])
    last_row += 1
    headers_2 = [["조합명", "투자기간 이내", "잔여재원여부", "투자의무비율 충족 필요", "주∙특수목적 충족 필요", "주∙특수목적 해당여부", "판단"]]
    sheet.update(f'A{last_row}:G{last_row}', headers_2)

    last_row += 1
    basic_info_df = get_basic_info_from_spreadsheet(file_id)
    second_filterd_df = basic_info_df[basic_info_df[8] == 'O']

    for _, row in second_filterd_df.iterrows():
        # 0번째 열과 9번째 열 값 추출
        data_to_add = [[row[0], row[8], '', '']]
        sheet.update(f'A{last_row}:D{last_row}', data_to_add)  # 열 A와 B에 각각 추가
        last_row += 1  # 다음 행으로 이동

    last_row += 1

    # 3번째 표 구성 (제목 및 헤더)
    sheet.update(f'A{last_row}', [['(3) 핵심운용인력 우선 배정 가능']])
    last_row += 1
    headers_3 = [["조합명", "투자기간 이내", "잔여재원여부", "핵심운용인력 우선 배정"]]
    sheet.update(f'A{last_row}:D{last_row}', headers_3)

    last_row += 1
    basic_info_df = get_basic_info_from_spreadsheet(file_id)
    second_filterd_df = basic_info_df[basic_info_df[8] == 'O']

    # '(2) 발굴자' 셀을 찾고 우측 셀 값을 조회
    try:
        finder_cell = sheet.find('(2) 발굴자')
        # 우측 셀 값부터 빈칸이 나올 때까지 탐색
        finder_right_values = []
        col = finder_cell.col + 1  # 첫 번째 우측 셀부터 시작
        while True:
            cell_value = sheet.cell(finder_cell.row, col).value  # 현재 셀 값 가져오기
            if not cell_value or cell_value.strip() == '':  # 빈칸이면 종료
                break
            finder_right_values.append(cell_value)  # 값을 리스트에 추가
            col += 1  # 다음 열로 이동

        # print(f'finder_right_values : {finder_right_values}')

        # 각 행에 대해 조건에 따라 값을 업데이트
        for _, row in second_filterd_df.iterrows():
            # 대표펀드 매니저와 핵심운용인력에 존재 여부 확인
            fund_manager = row[6]  # 실제 컬럼명으로 변경 필요
            core_staff = row[7]  # 실제 컬럼명으로 변경 필요
            
            # 각각의 값을 쉼표로 구분하여 분리
            fund_manager_list = fund_manager.split(',') if ',' in fund_manager else [fund_manager]
            core_staff_list = core_staff.split(',') if ',' in core_staff else [core_staff]

            # 두 리스트를 합쳐 전체 인원을 full_staff에 저장
            full_staff = fund_manager_list + core_staff_list
            full_staff = [name.strip() for name in full_staff]  # 각 이름의 앞뒤 공백 제거

            # finder_right_values 내 한명이라도 full_staff 내부에 존재한다면 'O' 아니면 'X'
            # finder_right_values 내 한명이라도 full_staff 내부에 존재한다면 'O' 아니면 'X'
            allocation_status = 'O' if any(name in full_staff for name in finder_right_values) else 'X'

            # 결과를 업데이트
            sheet.update(f'A{last_row}', [[row[0], row[8], '', allocation_status]])  # A열, B열, D열에 데이터 추가
            last_row += 1

    except gspread.exceptions.CellNotFound:
        print("셀 '(2) 발굴자'를 찾을 수 없습니다.")
    
    last_row += 2 #  Next pretention

def update_invest_summary(file_id, fund_name):
    # Google Sheets 파일에 인증하고 열기
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
    spreadsheet = gc.open_by_key(file_id)
    sheet = spreadsheet.sheet1  # 기본 워크시트 선택

    # 마지막으로 작성된 데이터 위치 확인
    data = sheet.get_all_values()
    last_row = len(data) + 2  # 마지막 데이터 이후 한 줄을 건너뜀
    
    # title
    sheet.update(f'A{last_row}', [['4. 투자 배분금액 및 근거']])
    last_row += 1
    sheet.update(f'A{last_row}', [['(1) 투자 배분금액']])
    last_row += 1

    headers_1 = [["투자조합명", "금번 투자금액"]]
    sheet.update(f'A{last_row}:B{last_row}', headers_1)
    last_row += 1

    # sheet에서 '(4) 투자예정금액' 셀을 탐색한다 - 해당 셀의 우측값을 조회한다
    # money에 결과를 저장한다 - '(4) 투자예정금액' 셀을 탐색하고 우측 값을 조회하여 money에 저장
    try:
        target_cell = sheet.find('(4) 투자예정금액')
        money = sheet.cell(target_cell.row, target_cell.col + 1).value  # 우측 셀 값 가져오기
        # DEBUGGING
        # print(f'fund_name: {fund_name}')
        # print(f"투자예정금액 우측 값: {money}")
        
    except gspread.exceptions.CellNotFound:
        print("셀 '(4) 투자예정금액'을 찾을 수 없습니다.")
        money = None  # 셀을 찾지 못했을 때 None으로 설정

    sheet.update(f'A{last_row}:B{last_row}', [[fund_name, money]])

    last_row = last_row + 2
    sheet.update(f'A{last_row}', [['(2) 배분 사유']])
    last_row = last_row + 1
    sheet.update(f'A{last_row}', [['조합 담당자가 사유 쓰기']])
    last_row = last_row + 1
    sheet.update(f'A{last_row}', [['EX) 투자 가능한 조합 중 주목적에 해당하며 재원이 있는 조합에 배정']])
    last_row = last_row + 1
    sheet.update(f'A{last_row}', [['투자 가능한 조합 중 기투자한 조합에 우선 배정']])
    last_row = last_row + 1
    sheet.update(f'A{last_row}', [['투자 가능한 조합 중 주목적, 특수목적 비율 충족이 필요한 조합에 배정']])
    last_row = last_row + 1
    sheet.update(f'A{last_row}', [['투자 가능한 조합 중 핵심운용인력이 발굴자로 있는 조합에 우선 배정']])

    return money

def get_all_data_from_spreadsheet(spreadsheeet_id):
    # 스프레드시트 ID 및 데이터 가져오기
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(spreadsheeet_id)
    worksheet = spreadsheet.sheet1  # 첫 번째 시트

    # 전체 데이터를 가져와 DataFrame으로 변환
    data = worksheet.get_all_values()
    df = pd.DataFrame(data)

    return df

def create_basic_info_table():
    db_7_spreadsheet_info = get_all_data_from_spreadsheet(config.db_7_id)

    NAME = 1   
    STATUS = 15 
    COMMITED_AMOUNT_INFO = 4
    ESTABLISHMENT_DATE = 10
    INVEST_EXPIRATION_DATE = 14
    DURATION_EXPIRATION_DATE = 11
    LEAD_FUND_MANAGER = 16 
    KEY_OPERATING_PERSONNEL = 17

    # db_7_spreadsheet_info 정보 중 '상태' 가 청산이 아닌 데이터만 조회합니다
    filtered_data = db_7_spreadsheet_info[db_7_spreadsheet_info[STATUS] != '청산']

    # 투자재원 채우기
    # filtered_data의 조합명과 비교, 일치하면 투자가능재원 값을 넣고 아니면 '-'로 메꾸기
    compared_file_id = config.investment_resources_overview_id
    client = gspread.authorize(credentials)
    # 스프레드시트 불러오기
    spreadsheet = client.open_by_key(compared_file_id)

    # '실시간'이라는 시트 가져오기
    sheet = spreadsheet.worksheet('실시간')

    # 3번째 행과 23번째 행 데이터 가져오기
    row_3 = sheet.row_values(3)
    row_21 = sheet.row_values(23)

    # 3번째 행의 데이터에서 공백을 제외한 요소를 리스트로 저장
    row_3_filtered = [item for item in row_3 if item.strip() != '']
    # 23번째 행에서 '투자집행가능'으로 시작하지 않는 요소와 공백을 제외한 요소를 리스트로 저장
    row_23_filtered = [item for item in row_21 if item.strip() != '' and not item.startswith('투자집행가능')]

    # row_3_filtered와 row_23_filtered 각각의 요소를 쌍으로 묶기
    paired_data = list(zip(row_3_filtered, row_23_filtered))
    # print(paired_data)

    result_list = []
    for name in row_3_filtered:
        matched = False
        for pair in paired_data:
            if name == pair[0]:
                result_list.append(pair[1])
                matched = True
                break
        if not matched:
            result_list.append('-')

    table_data = pd.DataFrame({
        '조합명': filtered_data[NAME],  # 예: 조합명을 0번 인덱스로 설정
        '결성금액': filtered_data[COMMITED_AMOUNT_INFO],
        '투자재원': None,  # 필요에 따라 수정
        '결성일': filtered_data[ESTABLISHMENT_DATE],
        '투자기간 만료일': filtered_data[INVEST_EXPIRATION_DATE],
        '존속기간 만료일': filtered_data[DURATION_EXPIRATION_DATE],
        '대표펀드 매니저': filtered_data[LEAD_FUND_MANAGER],
        '핵심운용인력': filtered_data[KEY_OPERATING_PERSONNEL],
        '투자기간': None  # 필요에 따라 수정
    })

    return table_data

def create_invest_Target_Company_information(db_1_df, db_4_df):
    # db_1_df에서 회사명 조회
    company_names = db_1_df['회사명'].iloc[0] if '회사명' in db_1_df.columns else None
    # print("회사명:", company_names)
    
    # db_4_df에서 발굴자1, 발굴자2, 발굴자3 조회
    discoverer1 = db_4_df['발굴자1'].iloc[0] if '발굴자1' in db_4_df.columns else None
    discoverer2 = db_4_df['발굴자2'].iloc[0] if '발굴자2' in db_4_df.columns else None
    discoverer3 = db_4_df['발굴자3'].iloc[0] if '발굴자3' in db_4_df.columns else None
    
    # db_4_df에서 'INV ID' 조회, 마지막 문자가 '1'이면 'O', 아니면 'X'
    if 'INV ID' in db_4_df.columns:
        INV_STATUS= db_4_df['INV ID'].apply(lambda x: 'X' if str(x).endswith('1') else 'O').iloc[0]
    
    # db_4_df에서 투자금액(원화) 조회
    investment_amount = db_4_df['투자금액(원화)'].iloc[0] if '투자금액(원화)' in db_4_df.columns else None
    
    # db_4_df에서 투자 납입일 조회
    investment_date = db_4_df['투자 납입일'].iloc[0] if '투자 납입일' in db_4_df.columns else None

    result_df = pd.DataFrame({
        '회사명': [company_names],
        '발굴자1': [discoverer1],
        '발굴자2': [discoverer2],
        '발굴자3': [discoverer3],
        'INV ID 상태': [INV_STATUS],
        '투자금액(원화)': [investment_amount],
        '투자 납입일': [investment_date]
    })

    return result_df

def preprocess_data(str_number):
    numeric_value = int(str_number.replace(',', ''))
    return int(numeric_value)

def preprocess_paired_data(paired_data):
        processed_data = []
        for name, value in paired_data:
            # 쉼표 제거, 음수 처리 포함
            try:
                numeric_value = int(value.replace(',', ''))
                numeric_value = int(numeric_value)
            except ValueError:
                numeric_value = 0  # 변환 실패 시 기본값
            processed_data.append((name, numeric_value))
        return processed_data

def append_A(paired_data, future_invest, file_id):
   
    client = gspread.authorize(credentials)
    # 스프레드시트 열기
    spreadsheet = client.open_by_key(file_id)
    worksheet = spreadsheet.sheet1  # 첫 번째 시트 가져오기

    # A열 데이터 가져오기
    column_A = worksheet.col_values(1)  # A열의 모든 데이터 가져오기

    # 데이터의 시작점 찾기
    start_row = None
    for i, cell_value in enumerate(column_A, start=1):
        if "(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부" in cell_value:
            start_row = i + 2  # 해당 위치의 +2 행부터 작업 시작
            break

    if start_row is None:
        raise ValueError("지정된 기준 문구를 찾을 수 없습니다.")
    
    # paired_data preprocessing
    # 예시 : [('카카오-신한 제1호 트나이트 투자조합', '187,026,695'), ('카카오 그로스해킹 펀드', '-879,743,279'), ('카카오 코파일럿 제1호 펀드', '501,399,532'), ('카카오 코파일럿 제2호 펀드', '7,434,649,829')]
    # 2번째 요소가 문자열인데 이를 숫자로 변환해야 함. 음수도 존재하는 상황임.
    # 숫자로 변환하는 코드가 필요
    processed_paired_data = preprocess_paired_data(paired_data)
    processed_future_invest = preprocess_data(future_invest)

    # 지정된 범위 내 데이터를 처리
    updates = []  # 업데이트할 데이터 저장
    row = start_row
    while True:
        # A열 값 가져오기
        cell_A_value = worksheet.cell(row, 1).value
        if not cell_A_value:  # 빈칸이 나오면 중단
            break

        # paired_data와 비교
        matched_pair = next((pair for pair in processed_paired_data if pair[0] == cell_A_value), None)
        if matched_pair:
            # 두 번째 요소와 future_invest 비교
            second_value = matched_pair[1]
            result = "O" if processed_future_invest <= second_value else "X"
        else:
            result = "-"

        # C열 업데이트 데이터 추가
        updates.append([result])
        row += 1  # 다음 행으로 이동

    # 업데이트를 한 번의 호출로 처리
    end_row = start_row + len(updates) - 1
    worksheet.update(f"C{start_row}:C{end_row}", updates)

    # file_id spreadsheet1 열어서 A열 값에 '(3) 핵심운용인력 우선 배정 가능'인 행 데이터 위치 구하기
    # 데이터 위치 + 2 의 C열에도 updates를 반영하려고 한다
    # updates 내 요소 개수만큼 행 데이터의 C열을 업데이트 한다 (요소 개수와 범위에 해당되는 행 데이터의 개수가 동일하다는 말)
    # '(3) 핵심운용인력 우선 배정 가능'을 찾기
    start_row_3 = next(
        (i + 2 for i, val in enumerate(column_A, start=1)
         if '(3) 핵심운용인력 우선 배정 가능' in val),
        None
    )
    if start_row_3 is None:
        raise ValueError("지정된 텍스트 '(3) 핵심운용인력 우선 배정 가능'을 찾을 수 없습니다.")

    # 해당 위치 +2부터 C열 업데이트
    updates_3 = []  # 새로운 업데이트 리스트
    row_3 = start_row_3
    for i in range(len(updates)):
        # C열 업데이트
        updates_3.append([updates[i][0]])

    # C열 업데이트 적용
    end_row_3 = start_row_3 + len(updates_3) - 1
    worksheet.update(f"C{start_row_3}:C{end_row_3}", updates_3)

def append_B(file_id, inv_id):
    # Google Sheets API 클라이언트 인증
    client = gspread.authorize(credentials)
    
    # 첫 번째 스프레드시트: 투자 재원 계산
    spreadsheet_1 = client.open_by_key(config.investment_present_table_id)
    worksheet_1 = spreadsheet_1.worksheet('통합')
    
    # A열, W열, AE열, AG열 데이터 가져오기
    column_A_1 = worksheet_1.col_values(1)  # A열 데이터
    column_W = worksheet_1.col_values(23)  # W열 데이터
    column_AE = worksheet_1.col_values(31)  # AE열 데이터
    column_AG = worksheet_1.col_values(33)  # AG열 데이터

    # 문자열 정리 함수
    def clean_number(value):
        try:
            return float(value.strip().replace(',', '').replace(' ', ''))  # 쉼표와 공백 제거 후 float 변환
        except (ValueError, AttributeError):
            return None  # 변환 실패 시 None 반환
        
    for i, value in enumerate(column_A_1):
        if value == inv_id:
            worksheet_name = worksheet_1.cell(i + 1, 3).value  # 해당 행의 C열 값 확인
            break
    else:
        raise ValueError(f"inv_id {inv_id} not found in column A.")

    # config.investment_present_table_id 내 스프레드시트 중에 이름이 worksheet_name인 워크시트의 전체 데이터를 조회
    worksheet_target = spreadsheet_1.worksheet(worksheet_name)
    rows = worksheet_target.get_all_values()

    # 데이터를 조회할 때 W열 값이 '여'인 행 데이터들만 조회
    filtered_rows = [
        row for row in rows if len(row) >= 23 and row[22] == '여'
    ]

    # 조회한 데이터들 중 A열 값이 inv_id인 행 데이터 위치를 확인하고 해당 위치 전까지의 행 데이터들만 남긴다. 이 때 순서를 유지해서 행 데이터를 남겨야 해.
    # 남은 행 데이터들에 대해서 각 행 데이터의 AE열값과 AG열 값을 곱함
    # 곱합 값들을 누적해서 합해서 invest_verified_sum 변수에 저장한다
    # config.investment_present_table_id 내 스프레드시트 중에 이름이 worksheet_name인 워크시트의 전체 데이터를 조회
    worksheet_target = spreadsheet_1.worksheet(worksheet_name)
    rows = worksheet_target.get_all_values()

    # 데이터를 조회할 때 W열 값이 '여'인 행 데이터들만 필터링
    filtered_rows = [
        row for row in rows if len(row) >= 23 and row[22] == '여'
    ]

    # 조회한 데이터들 중 A열 값이 inv_id인 행 데이터 위치를 확인
    for idx, row in enumerate(filtered_rows):
        if len(row) >= 1 and row[0] == inv_id:
            target_index = idx
            break
    else:
        target_index = len(filtered_rows)  # inv_id가 없으면 모든 데이터를 포함

    # 해당 위치 전까지의 데이터만 남김 (순서를 유지)
    remaining_rows = filtered_rows[:target_index]

    # 남은 행 데이터들에 대해서 각 행 데이터의 AE열 값과 AG열 값을 곱한 값을 누적
    invest_verified_sum = 0
    for row in remaining_rows:
        if len(row) >= 33:  # 데이터의 AE, AG 열이 존재하는지 확인
            ae_value = clean_number(row[30])  # AE열 값
            ag_value = clean_number(row[32])  # AG열 값
            if ae_value is not None and ag_value is not None:
                invest_verified_sum += ae_value * ag_value

    # 두 번째 스프레드시트: 조합명 - 투자약정금액 쌍 가져오기
    spreadsheet_2 = client.open_by_key(file_id)
    worksheet_2 = spreadsheet_2.sheet1  # 첫 번째 시트 가져오기
    column_A_2 = worksheet_2.col_values(1)
    column_B = worksheet_2.col_values(2)

    # '조합명' 위치 찾기
    start_row_2 = next(
        (i + 1 for i, val in enumerate(column_A_2, start=1) if val == '조합명'),
        None
    )

    if start_row_2 is None:
        raise ValueError("'조합명'을 찾을 수 없습니다.")

    table_1_pairs = []  # 결과를 저장할 리스트
    for row in range(start_row_2 -1, len(column_A_2)):
        a_value = column_A_2[row] if row < len(column_A_2) else None
        b_value = column_B[row] if row < len(column_B) else None

        # 빈칸을 만나면 중단
        if not a_value:
            break

        # 데이터가 유효한 경우 쌍으로 묶어 저장
        if a_value and b_value:
            table_1_pairs.append((a_value, b_value))
    
    # file_id의 spreadsheet를 열고 A열의 데이터 값에서 '(3) 핵심운용인력 우선 배정 가능'인 행 데이터의 위치를 탐색한다
    # 해당 행 데이터의 위치 +2 부터 빈칸을 만나기전까지 범위를 파악한다
    # 범위 내의 A열의 데이터 값들을 comparing_keys에 리스트 형식으로 저장한다
    start_row_3 = next(
        (i + 1 for i, val in enumerate(column_A_2, start=1)
         if val == '(3) 핵심운용인력 우선 배정 가능'),
        None
    )
    if start_row_3 is None:
        raise ValueError("지정된 텍스트 '(3) 핵심운용인력 우선 배정 가능'을 찾을 수 없습니다.")

    # 범위 내 A열 데이터 수집
    comparing_keys = []
    for row in range(start_row_3, len(column_A_2)):
        a_value = column_A_2[row] if row < len(column_A_2) else None
        # 빈칸을 만나면 중단
        if not a_value:
            break

        comparing_keys.append(a_value)

    # 조건 비교 및 결과 리스트 생성
    results = []
    for key in comparing_keys:
        # table_1_pairs에서 매칭된 값 찾기
        match = next((pair[1] for pair in table_1_pairs if pair[0] == key), None)
        if match is not None:
            # 비교: 20% 이상인지 확인
            # minus value checking!!!!!
            match = float(match.replace(',', '').strip())
            result = 'X' if invest_verified_sum >= (match * 0.2) else 'O'
            results.append(result)

    # file_id의 spreadsheet를 열고 A열의 데이터 값에서 '(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부'인 행 데이터 위치 찾기
    # 해당 행 데이터의 위치 + 2 부터 빈칸 전까지의 범위 구하기
    # 해당 범위의 D열에 result 내 요소를 순서대로 업데이트 하기 
    # '(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부' 텍스트 찾기
    start_row_4 = next(
        (i + 2 for i, val in enumerate(column_A_2, start=1)
         if '(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부' in val),
        None
    )
    if start_row_4 is None:
        raise ValueError("지정된 텍스트 '(2) 투자의무비율 & 주목적/특수목적비율 충족 필요 조합 우선 배정 여부'를 찾을 수 없습니다.")

    # 지정된 범위 내 C열에 results 리스트 값을 업데이트
    updates = []
    for i, result in enumerate(results):
        updates.append([result])  # 각 결과를 C열에 추가

    end_row_4 = start_row_4 + len(updates) - 1
    worksheet_2.update(f"D{start_row_4}:D{end_row_4}", updates)

def style_spreadsheet(file_id):
        client = gspread.authorize(credentials)
        # 파일을 가져오기
        spreadsheet = client.open_by_key(file_id)
        worksheet = spreadsheet.sheet1  # 기본적으로 첫 번째 시트를 선택

        # 열 너비를 데이터에 맞게 자동으로 조정하기
        for col in range(1, worksheet.col_count + 1):
            column_values = worksheet.col_values(col)
            max_length = max(len(str(value)) for value in column_values)
            worksheet.resize(rows=worksheet.row_count, cols=col)
            worksheet.format(f'{get_column_letter(col)}1', {'width': max_length * 1.2})

def create_investmentTable(db_1_df, db_4_df, db_7_df, kv_id, inv_id):
    fund_name = db_7_df['펀드명'].iloc[0]

    basic_df = create_basic_info_table()
    invest_df = create_invest_Target_Company_information(db_1_df, db_4_df)

    company_name = invest_df['회사명'].iloc[0]
    file_name = '투자재원배분표_' + company_name

    if check_folder_access(folder_id):
        if not is_file_exists_in_directory(folder_id, file_name):
            create_spreadsheet(folder_id, file_name)
            time.sleep(5)
        else:
            print(f"The file '{file_name}' already exists in the folder.")
    else:
        print("Please verify folder access permissions.")
    
    file_id = get_file_id_in_folder(folder_id, file_name)

    update_basic_with_dataframes(file_id, basic_df)
    update_invest_with_dataframes(file_id, invest_df)

    update_invest_date(file_id, basic_df, invest_df) # 투자 기간 업데이트하기
    paired_data = update_invest_src(file_id, basic_df) # 투자 재원 업데이트하기 
    update_invest_distribution(file_id, fund_name, inv_id)
    future_invest = update_invest_summary(file_id, fund_name)

    append_A(paired_data, future_invest, file_id) # 잔여재원여부 업데이트
    append_B(file_id, inv_id) # 투자의무 비율 충족 여부 

    # 스타일 변환하기 
    # style_spreadsheet(file_id)